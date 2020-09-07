import openpyxl
path = 'omknew.xlsx'                # path of file which want to read
wb = openpyxl.load_workbook(path)   # load spreadsheet to wb
sheet = wb.active                   # active workbook as sheet

# wb.create_sheet(index=1, title='second sheet')   # creating sheets in excel with title to it

# use for loop to read all values

cell = sheet.cell(row=1, column=1)  # giving values of rows and column
# cell2 = sheet['A2']               # another way to give values of rows and column
print(cell.value)                   # get values for rows and columns
cell.value = 82                     # to save values into spreadsheet

wb.save('after change.xlsx')        # after change values save file
