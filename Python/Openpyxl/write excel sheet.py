# import os
# os.getcwd()              # get current working directory before save files
# os.ch.dir("c:/myfiles")  # if want to change file location before save

from openpyxl import Workbook
import datetime

wb = Workbook()       # open workbook as wb

ws = wb.active        # get active wordsheet from workbook
ws['A1'] = 42         # saving 42 value to A1 position in spreadsheet
ws.append([1, 2, 3])  # saving list of 1, 2, 3 at new row of spreadsheet
ws['A2'] = datetime.datetime.now()  # saving date at A2 position in sheet

wb.save('omknew.xlsx')  # saving file as omknew.xlsx
